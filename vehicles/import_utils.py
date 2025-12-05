import csv
from io import TextIOWrapper
from typing import Any, Dict, List, Tuple


def read_csv(file) -> List[Dict[str, Any]]:
    wrapper = TextIOWrapper(file, encoding="utf-8") if hasattr(file, "read") else file
    reader = csv.DictReader(wrapper)
    return [dict(row) for row in reader]


def read_excel(file) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    headers = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h) if h is not None else "" for h in row]
            continue
        data = {}
        for j, val in enumerate(row):
            key = headers[j] if j < len(headers) else f"col_{j}"
            data[key] = val
        rows.append(data)
    return rows


def read_rows(file, file_type: str) -> List[Dict[str, Any]]:
    if file_type.lower() in ["csv"]:
        return read_csv(file)
    if file_type.lower() in ["xls", "xlsx", "excel"]:
        return read_excel(file)
    return []


def map_row(mapping: Dict[str, str], row: Dict[str, Any]) -> Dict[str, Any]:
    data = {}
    for field, column in mapping.items():
        data[field] = row.get(column)
    return data


def normalize_vehicle_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    from .models import Vehicule

    plate = payload.get("plaque_immatriculation")
    if plate:
        payload["plaque_immatriculation"] = Vehicule.normalize_plate(str(plate))
    if payload.get("puissance_fiscale_cv") is not None:
        try:
            payload["puissance_fiscale_cv"] = int(payload["puissance_fiscale_cv"])
        except Exception:
            pass
    if payload.get("cylindree_cm3") is not None:
        try:
            payload["cylindree_cm3"] = int(payload["cylindree_cm3"])
        except Exception:
            pass
    return payload


def validate_vehicle_payload(payload: Dict[str, Any]) -> Tuple[bool, List[str]]:
    errors = []
    required = [
        "plaque_immatriculation",
        "marque",
        "source_energie",
        "date_premiere_circulation",
        "puissance_fiscale_cv",
    ]
    for f in required:
        if not payload.get(f):
            errors.append(f"{f} manquant")
    source = payload.get("source_energie")
    if source and source not in ["Essence", "Diesel", "Electrique", "Hybride"]:
        errors.append("source_energie invalide")
    return (len(errors) == 0, errors)
